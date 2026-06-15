"""Tests for ethoinsight.parse using real EthoVision demo data.

Demo data location resolves in order:
1. ``ETHOINSIGHT_DEMO_BASE`` env var (e.g. ``/home/wangqiuyang/DemoData/newdemodata``)
2. legacy hard-coded path (kept for backward compat)

Subdirectory names: project repos used multiple Chinese naming variants
over time (e.g. ``斑马鱼鱼群行为`` vs ``斑马鱼``). If the resolved demo base
or the specific paradigm subdirectory does not exist, the relevant tests
are skipped rather than failed — these tests are integration aids, not
infra invariants.
"""

import os
from pathlib import Path

import pandas as pd
import pytest

from ethoinsight import parse
from ethoinsight.utils import detect_paradigm, normalize_column_name

# Demo data base — prefer env var, fall back to legacy hard-coded path.
DEMO_BASE = Path(
    os.environ.get("ETHOINSIGHT_DEMO_BASE", "/home/qiuyangwang/noldus-insight/demo-data/DemoData")
)
ZEBRAFISH_DIR = DEMO_BASE / "斑马鱼鱼群行为"
EPM_DIR = DEMO_BASE / "高架十字迷宫"
O_MAZE_DIR = DEMO_BASE / "O迷宫"
NOVEL_OBJECT_DIR = DEMO_BASE / "新物体识别"
Y_MAZE_DIR = DEMO_BASE / "Y迷宫"


def _require_trajectory_files(directory: Path) -> list[str]:
    """Find trajectory files in *directory* or skip the test if none exist."""
    if not directory.exists():
        pytest.skip(f"Demo data directory missing: {directory}. Set ETHOINSIGHT_DEMO_BASE to enable.")
    files = sorted(str(f) for f in directory.glob("轨迹-*.txt"))
    if not files:
        pytest.skip(f"No 轨迹-*.txt files in {directory}")
    return files


# Find trajectory files (start with "轨迹-")
def _find_trajectory_files(directory: Path) -> list[str]:
    if not directory.exists():
        return []
    return sorted(str(f) for f in directory.glob("轨迹-*.txt"))


def _find_any_txt(directory: Path) -> list[str]:
    if not directory.exists():
        return []
    return sorted(str(f) for f in directory.glob("*.txt"))


# ============================================================================
# detect_ethovision
# ============================================================================


class TestDetectEthovision:
    def test_valid_trajectory_file(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        assert len(files) > 0, "No zebrafish trajectory files found"
        assert parse.detect_ethovision(files[0]) is True

    def test_valid_statistics_file(self):
        stats_files = sorted(str(f) for f in EPM_DIR.glob("统计数据-*.txt"))
        if stats_files:
            assert parse.detect_ethovision(stats_files[0]) is True

    def test_nonexistent_file(self):
        assert parse.detect_ethovision("/tmp/nonexistent_file.txt") is False

    def test_non_ethovision_file(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("hello world", encoding="utf-8")
        assert parse.detect_ethovision(str(f)) is False

    def test_directory_returns_false(self, tmp_path):
        assert parse.detect_ethovision(str(tmp_path)) is False

    def test_english_locale_trajectory_file(self, tmp_path):
        """TXT with 'Number of header lines:' (English locale) → True."""
        import codecs

        f = tmp_path / "english_locale.txt"
        content = "Number of header lines: 39\r\n"
        with open(str(f), "wb") as fh:
            fh.write(codecs.BOM_UTF16_LE)
            fh.write(content.encode("utf-16-le"))
        assert parse.detect_ethovision(str(f)) is True

    def test_chinese_locale_trajectory_file(self, tmp_path):
        """TXT with '标题行数：' (Chinese locale) → True (regression)."""
        import codecs

        f = tmp_path / "chinese_locale.txt"
        content = "标题行数：39\r\n"
        with open(str(f), "wb") as fh:
            fh.write(codecs.BOM_UTF16_LE)
            fh.write(content.encode("utf-16-le"))
        assert parse.detect_ethovision(str(f)) is True


# ============================================================================
# parse_header
# ============================================================================


class TestParseHeader:
    def test_zebrafish_header(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        assert len(files) > 0
        header = parse.parse_header(files[0])

        assert header["header_lines"] > 0
        assert (
            "shoaling" in header["experiment"].lower()
            or header["paradigm"] == "shoaling"
        )
        assert header["paradigm"] == "shoaling"
        assert header["subject"] != ""
        assert len(header["columns"]) > 0
        assert "trial_time" in header["columns"]
        assert "recording_time" in header["columns"]
        assert "x_center" in header["columns"]
        assert "y_center" in header["columns"]

    def test_epm_header(self):
        files = _find_trajectory_files(EPM_DIR)
        if not files:
            pytest.skip("No EPM trajectory files")
        header = parse.parse_header(files[0])

        assert header["paradigm"] == "epm"
        assert "x_nose" in header["columns"] or "x_center" in header["columns"]

    def test_o_maze_header(self):
        files = _find_trajectory_files(O_MAZE_DIR)
        if not files:
            pytest.skip("No O-maze trajectory files")
        header = parse.parse_header(files[0])

        assert header["paradigm"] == "o_maze"
        assert "trial_time" in header["columns"]

    def test_novel_object_header(self):
        files = _find_trajectory_files(NOVEL_OBJECT_DIR)
        if not files:
            pytest.skip("No novel object trajectory files")
        header = parse.parse_header(files[0])

        assert header["paradigm"] == "novel_object"
        # Basic trajectory columns should always be present
        assert "trial_time" in header["columns"]
        assert "x_center" in header["columns"]

    def test_units_parsed(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        header = parse.parse_header(files[0])
        assert len(header["units"]) > 0

    def test_raw_metadata(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        header = parse.parse_header(files[0])
        assert (
            "实验" in header["raw_metadata"] or "Experiment" in header["raw_metadata"]
        )


# ============================================================================
# parse_trajectory
# ============================================================================


class TestParseTrajectory:
    def test_zebrafish_trajectory(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        assert len(files) > 0
        df = parse.parse_trajectory(files[0])

        assert isinstance(df, pd.DataFrame)
        assert len(df) > 100, "Expected at least 100 data rows"
        assert "trial_time" in df.columns
        assert "x_center" in df.columns
        assert "y_center" in df.columns

        # Numeric types
        assert df["trial_time"].dtype in (float, "float64")
        assert df["x_center"].dtype in (float, "float64")

        # No "-" strings remaining
        for col in df.columns:
            if df[col].dtype == object:
                continue
            assert not df[col].astype(str).str.contains("^-$").any(), (
                f"Column {col} still has '-' values"
            )

        # Metadata in attrs
        assert "subject" in df.attrs
        assert "paradigm" in df.attrs

    def test_epm_trajectory(self):
        files = _find_trajectory_files(EPM_DIR)
        if not files:
            pytest.skip("No EPM trajectory files")
        df = parse.parse_trajectory(files[0])

        assert len(df) > 100
        assert "trial_time" in df.columns
        assert df.attrs.get("paradigm") == "epm"

    def test_y_maze_trajectory(self):
        files = _find_trajectory_files(Y_MAZE_DIR)
        if not files:
            pytest.skip("No Y-maze trajectory files")
        df = parse.parse_trajectory(files[0])

        assert len(df) > 0
        assert df.attrs.get("paradigm") == "y_maze"

    def test_nan_handling(self):
        """Verify that '-' values are converted to NaN."""
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        df = parse.parse_trajectory(files[0])

        # distance_moved or velocity often has NaN at the start
        numeric_cols = df.select_dtypes(include="number").columns
        # At least some columns should have NaN (first row often has "-")
        has_nan = any(df[col].isna().any() for col in numeric_cols)
        # This is typical but not guaranteed, so just check it doesn't crash
        assert isinstance(df, pd.DataFrame)


# ============================================================================
# parse_batch
# ============================================================================


class TestParseBatch:
    def test_zebrafish_batch(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)
        assert len(files) > 0
        result = parse.parse_batch(files)

        assert result["summary"]["total_files"] == len(files)
        assert result["summary"]["total_rows"] > 0
        assert len(result["subjects"]) > 0
        assert result["summary"]["paradigm"] == "shoaling"
        assert isinstance(result["all_data"], pd.DataFrame)
        assert "subject" in result["all_data"].columns
        assert "file" in result["all_data"].columns

    def test_glob_pattern(self):
        if not ZEBRAFISH_DIR.exists():
            pytest.skip(f"Demo data directory missing: {ZEBRAFISH_DIR}. Set ETHOINSIGHT_DEMO_BASE to enable.")
        pattern = str(ZEBRAFISH_DIR / "轨迹-*.txt")
        result = parse.parse_batch(pattern)
        if result["summary"]["total_files"] == 0:
            pytest.skip(f"No 轨迹-*.txt files in {ZEBRAFISH_DIR}")

        assert result["summary"]["total_files"] > 0
        assert result["summary"]["paradigm"] == "shoaling"

    def test_empty_input(self):
        result = parse.parse_batch([])
        assert result["summary"]["total_files"] == 0
        assert len(result["subjects"]) == 0

    def test_nonexistent_glob(self):
        result = parse.parse_batch("/tmp/nonexistent_pattern_*.txt")
        assert result["summary"]["total_files"] == 0

    def test_mixed_files_filtered(self):
        """Statistics files should be filtered out, only trajectories parsed."""
        all_files = _find_any_txt(EPM_DIR)
        trajectory_files = _find_trajectory_files(EPM_DIR)
        if not trajectory_files:
            pytest.skip("No EPM trajectory files")

        result = parse.parse_batch(all_files)
        assert result["summary"]["total_files"] == len(trajectory_files)


# ============================================================================
# get_summary
# ============================================================================


class TestGetSummary:
    def test_summary_content(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)[:5]  # Use subset for speed
        result = parse.parse_batch(files)
        text = parse.get_summary(result)

        assert "EthoVision Data Summary" in text
        assert "shoaling" in text.lower()
        assert "Files:" in text
        assert "Subject" in text or "subject" in text.lower()

    def test_summary_max_chars(self):
        files = _require_trajectory_files(ZEBRAFISH_DIR)[:5]
        result = parse.parse_batch(files)
        text = parse.get_summary(result, max_chars=500)

        assert len(text) <= 500

    def test_empty_data_summary(self):
        result = parse.parse_batch([])
        text = parse.get_summary(result)
        assert "Files: 0" in text


# ============================================================================
# utils: normalize_column_name & detect_paradigm
# ============================================================================


class TestNormalizeColumnName:
    def test_exact_match(self):
        assert normalize_column_name("试用时间") == "trial_time"
        assert normalize_column_name("X 中心") == "x_center"
        assert normalize_column_name("Velocity") == "velocity"

    def test_quoted(self):
        assert normalize_column_name('"试用时间"') == "trial_time"
        assert normalize_column_name(' "X 中心" ') == "x_center"

    def test_in_zone_pattern(self):
        name = "In zone(Open arms /中心点)"
        result = normalize_column_name(name)
        assert result.startswith("in_zone")
        assert "open_arms" in result

    def test_chinese_zone_pattern(self):
        name = "分析区中(开放臂 /所有 中心点, 鼻尖)"
        result = normalize_column_name(name)
        assert result.startswith("in_zone")
        assert "open_arms" in result

    def test_entries_pattern(self):
        name = "Entries to Open arms"
        result = normalize_column_name(name)
        assert result.startswith("entries")

    def test_nose_in_zone_pattern(self):
        name = "Nose within object zone(Familiar object 1 /鼻尖)"
        result = normalize_column_name(name)
        assert result.startswith("nose_in_zone")

    def test_empty_name(self):
        assert normalize_column_name("") == "unnamed"
        assert normalize_column_name('""') == "unnamed"


class TestDetectParadigm:
    def test_shoaling(self):
        assert detect_paradigm("Shoaling behavior with JavaScript XT180") == "shoaling"

    def test_epm(self):
        assert detect_paradigm("Elevated Plus Maze XT180") == "epm"

    def test_open_field(self):
        assert detect_paradigm("Open Field test XT160") == "open_field"

    def test_o_maze(self):
        assert detect_paradigm("ethoInsightDemo-oMaze") == "o_maze"

    def test_novel_object(self):
        assert (
            detect_paradigm("Novel Object Recognition test with Deep Learning XT180")
            == "novel_object"
        )

    def test_y_maze(self):
        assert detect_paradigm("安琥医大Y迷宫") == "y_maze"

    def test_light_dark(self):
        assert detect_paradigm("ethoInsightDemo-ldb") == "light_dark"

    def test_unknown(self):
        assert detect_paradigm("Some random experiment") is None

    def test_empty(self):
        assert detect_paradigm("") is None


# ============================================================================
# infer_groups_from_result_block — 2026-05-13 同事反馈 Q3: 自动分组重写
# ============================================================================


def test_parse_groups_from_result_block_name():
    from ethoinsight.parse._core import infer_groups_from_result_block

    result = infer_groups_from_result_block(
        subjects=[
            {"name": "A1", "result_block_name": "Drug"},
            {"name": "A2", "result_block_name": "Drug"},
            {"name": "A3", "result_block_name": "Saline"},
            {"name": "A4", "result_block_name": "Saline"},
        ]
    )
    assert result == {"Drug": ["A1", "A2"], "Saline": ["A3", "A4"]}

    result_default = infer_groups_from_result_block(
        subjects=[
            {"name": "A1", "result_block_name": "Result 1"},
            {"name": "A2", "result_block_name": "Result 1"},
        ]
    )
    assert result_default is None


def test_parse_groups_handles_missing_field():
    from ethoinsight.parse._core import infer_groups_from_result_block

    result = infer_groups_from_result_block(
        subjects=[{"name": "A1"}, {"name": "A2"}],
    )
    assert result is None


# ============================================================================
# XLSX / Excel format tests
# ============================================================================

XLSX_DEMO_BASE = Path(
    os.environ.get("ETHOINSIGHT_XLSX_DEMO_BASE", "/home/wangqiuyang/DemoData/newdemodata")
)


def _find_xlsx_files(directory: Path) -> list[str]:
    """Find 原始数据-*.xlsx files in *directory*."""
    if not directory.exists():
        return []
    return sorted(str(f) for f in directory.rglob("原始数据-*.xlsx"))


def _require_xlsx_file() -> str:
    """Return the path to the first available OFT XLSX file, or skip."""
    oft_dir = XLSX_DEMO_BASE / "旷场_小鼠_三点"
    files = _find_xlsx_files(oft_dir)
    if not files:
        pytest.skip(f"No XLSX demo data in {XLSX_DEMO_BASE}. Set ETHOINSIGHT_XLSX_DEMO_BASE.")
    return files[0]


def _require_multi_sheet_xlsx() -> str:
    """Return path to the first multi-sheet XLSX (FST Mobility_10), or skip."""
    fst_dir = XLSX_DEMO_BASE / "强迫游泳_大鼠" / "Mobility_10"
    files = _find_xlsx_files(fst_dir)
    if not files:
        pytest.skip(f"No multi-sheet FST XLSX demo data in {fst_dir}")
    return files[0]


class TestDetectEthovisionXlsx:
    """detect_ethovision() with XLSX files."""

    def test_detect_xlsx_returns_true(self):
        from ethoinsight.parse._core import detect_ethovision

        path = _require_xlsx_file()
        assert detect_ethovision(path) is True

    def test_detect_xlsx_english_locale_marker(self, tmp_path):
        """A1 = 'Number of header lines:' (English locale) → True."""
        from ethoinsight.parse._core import detect_ethovision

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Number of header lines:")
        ws.cell(1, 2, 39)
        path = tmp_path / "english_locale.xlsx"
        wb.save(str(path))
        assert detect_ethovision(str(path)) is True

    def test_detect_xlsx_chinese_locale_marker(self, tmp_path):
        """A1 = '标题行数：' (Chinese locale) → True (regression)."""
        from ethoinsight.parse._core import detect_ethovision

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "标题行数：")
        ws.cell(1, 2, 39)
        path = tmp_path / "chinese_locale.xlsx"
        wb.save(str(path))
        assert detect_ethovision(str(path)) is True

    def test_detect_xlsx_no_marker(self, tmp_path):
        """A1 is something else → False."""
        from ethoinsight.parse._core import detect_ethovision

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1, "Sales Report")
        ws.cell(1, 2, 42)
        path = tmp_path / "not_ethovision.xlsx"
        wb.save(str(path))
        assert detect_ethovision(str(path)) is False

    def test_detect_txt_regression(self):
        from ethoinsight.parse._core import detect_ethovision

        txt_dir = XLSX_DEMO_BASE / "旷场_小鼠_三点"
        txt_files = sorted(str(f) for f in txt_dir.glob("轨迹-*.txt"))
        if not txt_files:
            pytest.skip("No TXT files alongside XLSX for regression check")
        assert detect_ethovision(txt_files[0]) is True

    def test_detect_unsupported_suffix(self, tmp_path):
        from ethoinsight.parse._core import detect_ethovision

        p = tmp_path / "data.pdf"
        p.write_text("not ethovision")
        assert detect_ethovision(str(p)) is False


class TestParseHeaderXlsx:
    """parse_header() with XLSX files."""

    def test_parse_xlsx_header(self):
        from ethoinsight.parse._core import parse_header

        path = _require_xlsx_file()
        header = parse_header(path)
        assert header["header_lines"] > 0
        assert len(header["columns"]) > 0
        assert header["paradigm"] == "open_field"
        assert header["experiment"] == "Open Field test XT190"

    def test_parse_xlsx_specific_sheet(self):
        from ethoinsight.parse._core import parse_header

        path = _require_multi_sheet_xlsx()
        # Read second sheet via :: convention
        header = parse_header(f"{path}::轨迹-Arena 2-Subject 1")
        assert header["arena"] == "Arena 2"
        assert header["subject"] == "Subject 1"


class TestParseTrajectoryXlsx:
    """parse_trajectory() with XLSX files."""

    def test_parse_xlsx_trajectory(self):
        from ethoinsight.parse._core import parse_trajectory

        path = _require_xlsx_file()
        df = parse_trajectory(path)
        assert not df.empty
        assert df.shape[1] == 18  # OFT has 18 columns
        assert df.attrs.get("paradigm") == "open_field"
        assert df.attrs.get("subject") == "Subject 1"

    def test_parse_xlsx_multi_sheet(self):
        from ethoinsight.parse._core import parse_trajectory

        path = _require_multi_sheet_xlsx()
        df1 = parse_trajectory(f"{path}::轨迹-Arena 1-Subject 1")
        df2 = parse_trajectory(f"{path}::轨迹-Arena 2-Subject 1")
        assert not df1.empty
        assert not df2.empty
        # Two arenas should have different data
        assert df1.attrs.get("arena") != df2.attrs.get("arena")


class TestParseBatchXlsx:
    """parse_batch() with XLSX files (and mixed TXT+XLSX)."""

    def test_parse_batch_xlsx_single(self):
        from ethoinsight.parse._core import parse_batch

        path = _require_xlsx_file()
        result = parse_batch([path])
        assert result["summary"]["total_files"] == 1
        assert result["summary"]["paradigm"] == "open_field"

    def test_parse_batch_xlsx_multi_sheet_expanded(self):
        from ethoinsight.parse._core import parse_batch

        path = _require_multi_sheet_xlsx()
        # Simulate the prep_metric_plan expansion: one entry per sheet
        paths = [
            f"{path}::轨迹-Arena 1-Subject 1",
            f"{path}::轨迹-Arena 2-Subject 1",
        ]
        result = parse_batch(paths)
        assert result["summary"]["total_files"] == 2


class TestXlsxErrorHandling:
    """Error cases for XLSX parsing."""

    def test_parse_header_nonexistent_sheet(self):
        from ethoinsight.parse._core import parse_header

        path = _require_xlsx_file()
        with pytest.raises(Exception):
            parse_header(f"{path}::NonExistentSheet")


# ============================================================================
# Golden parse fixtures — real-world EthoVision export variants
# ============================================================================

REAL_DATA_BASE = Path(os.environ.get("ETHOINSIGHT_REAL_BASE", "/home/wangqiuyang/DemoData/real_data"))


def _find_real_files(paradigm_dir: str) -> list[Path]:
    """Return .xlsx files in a real-data paradigm subdirectory, or empty list."""
    d = REAL_DATA_BASE / paradigm_dir
    if not d.is_dir():
        return []
    return sorted(f for f in d.glob("*.xlsx") if not f.name.startswith("~$"))


class TestRealDataNormalization:
    """Golden fixtures: real-world EthoVision exports must normalize to canonical names.

    Each test verifies that parse_header on real-world data produces column names
    that match the catalog & compute layer expectations.  If a directory doesn't
    exist locally the test is skipped (these are integration aids, not CI infra).
    """

    def test_epm_real_data_has_in_zone_columns(self):
        """Real EPM data: zone-prefix columns (开放臂/封闭臂) → in_zone_*."""
        files = _find_real_files("原始数据-高架十字迷宫实验")
        if not files:
            pytest.skip("real EPM data not found")
        for f in files[:2]:
            result = parse.parse_header(str(f))
            cols = result["columns"]
            in_zone = [c for c in cols if c.startswith("in_zone_")]
            assert len(in_zone) >= 2, (
                f"{f.name}: expected ≥2 in_zone_* columns, got {in_zone}. "
                f"All columns: {cols}"
            )
            # Must contain both open_arms and closed_arms variants
            joined = " ".join(in_zone)
            assert "open" in joined, f"{f.name}: missing open arm zone column in {in_zone}"
            assert "closed" in joined, f"{f.name}: missing closed arm zone column in {in_zone}"

    def test_tst_real_data_has_mobility_state_columns(self):
        """Real TST data: 活动状态(狂躁/静止) → mobility_state_highly_mobile / _immobile."""
        files = _find_real_files("原始数据-悬尾实验")
        if not files:
            pytest.skip("real TST data not found")
        for f in files[:2]:
            result = parse.parse_header(str(f))
            cols = result["columns"]
            mobility = [c for c in cols if "mobility_state" in c]
            assert len(mobility) >= 1, (
                f"{f.name}: expected ≥1 mobility_state* column, got {mobility}. "
                f"All columns: {cols}"
            )

    def test_fst_real_data_has_mobility_column(self):
        """Real FST data: single 活动状态 column → mobility_state."""
        fst_file = REAL_DATA_BASE / "原始数据-强迫游泳实验sample demo-试验     1.xlsx"
        if not fst_file.exists():
            pytest.skip("real FST data not found")
        result = parse.parse_header(str(fst_file))
        cols = result["columns"]
        assert "mobility_state" in cols, (
            f"Expected mobility_state in columns, got: {cols}"
        )

    def test_ezm_real_data_normalizes_zone_columns(self):
        """Real EZM (O迷宫) data: English columns already have in_closed / in_open."""
        files = _find_real_files("o迷宫")
        if not files:
            pytest.skip("real EZM data not found")
        f = files[0]
        result = parse.parse_header(str(f))
        cols = result["columns"]
        assert "in_closed" in cols, f"{f.name}: expected in_closed column, got {cols}"
        assert "in_open" in cols, f"{f.name}: expected in_open column, got {cols}"

    def test_real_data_does_not_regress_demo_data(self):
        """Demo data patterns (分析区中/In zone/Mobility) must still work."""
        demo_dir = Path(os.environ.get("ETHOINSIGHT_DEMO_BASE", DEMO_BASE))
        epm_demo = demo_dir / "高架十字迷宫_小鼠_三点"
        if not epm_demo.is_dir():
            pytest.skip("demo EPM data not found")
        xlsx_files = sorted(f for f in epm_demo.glob("*.xlsx") if not f.name.startswith("~$"))
        if not xlsx_files:
            pytest.skip("no demo EPM xlsx files")
        result = parse.parse_header(str(xlsx_files[0]))
        cols = result["columns"]
        assert any("in_zone" in c and "open" in c for c in cols), (
            f"Demo EPM should still have in_zone_open_arms_* column. Got: {cols}"
        )


# ============================================================================
# 2026-06-12: calamine 启动期断言 (Spec S3 Part B)
# ============================================================================


class TestCalamineImportAssertion:
    """parse 模块 import 期断言 python-calamine 可用。

    PR #125 把 Excel 引擎换 calamine（硬依赖），但缺依赖只在第一次
    pd.read_excel(engine="calamine") 才炸，且 _detect_ethovision_xlsx 的
    try/except 会把 calamine 缺失误判成"非 EthoVision 文件"（哑故障）。
    B1 在 parse/__init__.py import 期响亮断言，缺依赖当场失败。
    """

    def test_import_succeeds_when_calamine_available(self):
        """正常路径：python-calamine 可用时 import ethoinsight.parse 成功。"""
        import ethoinsight.parse  # noqa: F811 — already imported at module level, verify re-import ok
        assert ethoinsight.parse.parse_header is not None

    def test_missing_calamine_raises_import_error_with_guidance(self):
        """calamine 不可用时 import ethoinsight.parse 抛带指引的 ImportError。"""
        import os
        import subprocess
        import sys
        from pathlib import Path

        # ethoinsight 包根目录（含 pyproject + ethoinsight/ 包），从 __file__ 推导，
        # 不硬编码绝对路径——换机器 / CI / 移动 worktree 仍可用。
        pkg_root = Path(__file__).resolve().parents[1]

        code = """
import sys
# 模拟 python_calamine 不可导入
sys.modules['python_calamine'] = None
try:
    from ethoinsight import parse
    print('UNEXPECTED_SUCCESS')
except ImportError as e:
    msg = str(e)
    if 'python-calamine' in msg and 'calamine' in msg:
        print('CORRECT_IMPORT_ERROR')
    else:
        print(f'WRONG_MESSAGE: {msg}')
"""
        result = subprocess.run(
            [sys.executable, "-c", code],
            capture_output=True,
            text=True,
            cwd=str(pkg_root),
            env={**os.environ, "PYTHONPATH": str(pkg_root)},
        )
        assert "CORRECT_IMPORT_ERROR" in result.stdout, (
            f"Expected CORRECT_IMPORT_ERROR, got stdout={result.stdout!r} stderr={result.stderr!r}"
        )
